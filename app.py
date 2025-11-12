# save as replace_word_from_excel.py
# Requires: python-docx, openpyxl
# pip install python-docx openpyxl

import docx
import openpyxl
from tkinter import Tk, filedialog
from docx.oxml import OxmlElement
from docx.oxml.ns import qn
import os

def get_all_paragraph_elements_part(part_element):
    """Yield all w:p paragraph elements under a part element (body/header/footer),
       including those inside textboxes (w:txbxContent)."""
    for node in part_element.iter():
        # tag looks like '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'
        if node.tag.endswith('}p'):
            yield node

def get_text_nodes_from_p(p_elem):
    """Return list of w:t elements (text nodes) under a w:p paragraph element (in order)."""
    t_nodes = []
    for node in p_elem.iter():
        if node.tag.endswith('}t'):
            t_nodes.append(node)
    return t_nodes

def paragraph_has_underline(p_elem):
    """Return True if any run (w:r) under this paragraph has an underline element (w:u) or w:u value."""
    for r in p_elem.iter():
        if r.tag.endswith('}r'):  # run
            for rpr in r.iterfind('.//{}'.format(qn('w:rPr'))):
                # find underline element
                for u in rpr.iter():
                    if u.tag.endswith('}u'):
                        # if there's an underline element and it's not 'none'
                        val = u.get(qn('w:val'))
                        # If val is None it still means underline on default; treat as underlined
                        if val is None or val.lower() != 'none':
                            return True
    return False

def replace_in_part_xml(part_element, old, new):
    """Replace occurrences of old->new inside a document part element (body/header/footer),
       handling split runs and textboxes by working at w:p/w:t XML level."""
    replaced_count = 0
    for p in get_all_paragraph_elements_part(part_element):
        t_nodes = get_text_nodes_from_p(p)
        if not t_nodes:
            continue
        # join texts in order
        full = ''.join((t.text or '') for t in t_nodes)
        if old not in full:
            continue

        # decide uppercase if any run in this paragraph is underlined
        under = paragraph_has_underline(p)
        if under:
            replacement = full.replace(old, new.upper())
        else:
            replacement = full.replace(old, new)

        # write replacement back: set first t node to new full text, clear the others
        # this will collapse the paragraph's text into the first text node (practical and stable)
        t_nodes[0].text = replacement
        for t in t_nodes[1:]:
            t.text = ''

        replaced_count += 1
    return replaced_count

def replace_everywhere_doc(doc, old, new):
    total = 0
    # BODY
    total += replace_in_part_xml(doc.element.body, old, new)

    # HEADERS/FOOTERS - loop each section
    for section in doc.sections:
        try:
            hdr = section.header
            total += replace_in_part_xml(hdr._element, old, new)
        except Exception:
            # fallback if _element not available
            try:
                total += replace_in_part_xml(section.header.element, old, new)
            except Exception:
                pass
        try:
            ftr = section.footer
            total += replace_in_part_xml(ftr._element, old, new)
        except Exception:
            try:
                total += replace_in_part_xml(section.footer.element, old, new)
            except Exception:
                pass

    # Also try all 'parts' that might contain paragraphs (safety)
    # Some documents have additional parts; iterate over doc.part.rels to find possible header/footer parts
    try:
        for rel in doc.part.rels.values():
            part = getattr(rel._target, "element", None) or getattr(rel._target, "_element", None)
            if part is not None:
                # part may be header/footer/body; attempt replace there too
                total += replace_in_part_xml(part, old, new)
    except Exception:
        pass

    return total

def main():
    Tk().withdraw()
    print("Select Excel file (A2 down = old, B2 down = new)")
    excel_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xls")])
    if not excel_path:
        print("No excel selected. Exiting.")
        return

    print("Select Master Word (.docx) file")
    word_path = filedialog.askopenfilename(filetypes=[("Word files", "*.docx")])
    if not word_path:
        print("No Word file selected. Exiting.")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    doc = docx.Document(word_path)

    # collect replacements from A2 downward
    replacements = []
    for row in range(2, ws.max_row + 1):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        if a is None:
            continue
        if b is None:
            continue
        replacements.append((str(a), str(b)))

    if not replacements:
        print("No replacements found in A2:B... Exiting.")
        return

    overall_changes = 0
    for old, new in replacements:
        changed = replace_everywhere_doc(doc, old, new)
        print(f"Replaced occurrences of '{old}' → '{new}' in {changed} paragraph blocks (XML-level).")
        overall_changes += changed

    # save
    base, ext = os.path.splitext(word_path)
    out_path = base + "_UPDATED.docx"
    doc.save(out_path)
    print(f"\nDone. Total paragraph blocks changed: {overall_changes}")
    print("Output saved to:", out_path)
    print("If some replacements still didn't happen, tell me one example 'old' text that failed and I will adapt further.")

if __name__ == "__main__":
    main()
